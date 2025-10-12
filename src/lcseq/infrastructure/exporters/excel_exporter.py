"""
Excel exporter for analysis results.

Exports analysis results to Excel format with multiple sheets.
"""

import pandas as pd
from pathlib import Path
from typing import Dict
from ...application.dtos.analysis_response import AnalysisResponse


class ExcelExporter:
    """
    Export analysis results to Excel format.

    Creates Excel file with multiple sheets:
    - Compound Results
    - Validation Summary
    - Dataset Statistics

    Examples
    --------
    >>> exporter = ExcelExporter()
    >>> response = AnalysisResponse(...)
    >>> exporter.export(response, Path('results/analysis.xlsx'))
    """

    def export(
        self,
        response: AnalysisResponse,
        output_path: Path
    ) -> Path:
        """
        Export analysis results to Excel.

        Parameters
        ----------
        response : AnalysisResponse
            Analysis results to export
        output_path : Path
            Output file path (should end with .xlsx)

        Returns
        -------
        Path
            Path to created Excel file

        Examples
        --------
        >>> excel_path = exporter.export(response, Path('results/analysis.xlsx'))
        """
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Compound Results
            compound_data = [result.to_dict() for result in response.compound_results]
            df_compounds = pd.DataFrame(compound_data)
            df_compounds.to_excel(writer, sheet_name='Compound Results', index=False)

            # Sheet 2: Validation Summary
            summary_data = response.validation_summary.to_dict()
            df_summary = pd.DataFrame([summary_data])
            df_summary.to_excel(writer, sheet_name='Validation Summary', index=False)

            # Sheet 3: Dataset Statistics
            stats_data = response.dataset_stats
            df_stats = pd.DataFrame([stats_data])
            df_stats.to_excel(writer, sheet_name='Dataset Statistics', index=False)

            # Sheet 4: Processing Metadata
            metadata_data = response.processing_metadata
            df_metadata = pd.DataFrame([metadata_data])
            df_metadata.to_excel(writer, sheet_name='Processing Metadata', index=False)

        return output_path

    def export_with_formatting(
        self,
        response: AnalysisResponse,
        output_path: Path
    ) -> Path:
        """
        Export with Excel formatting and conditional formatting.

        Parameters
        ----------
        response : AnalysisResponse
            Analysis results
        output_path : Path
            Output file path

        Returns
        -------
        Path
            Path to formatted Excel file

        Notes
        -----
        Adds:
        - Color-coded validation status
        - Bold headers
        - Auto-sized columns
        - Frozen header row

        Examples
        --------
        >>> path = exporter.export_with_formatting(response, Path('results.xlsx'))
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        # First export normally
        self.export(response, output_path)

        # Load and format
        wb = load_workbook(output_path)

        # Format Compound Results sheet
        ws_compounds = wb['Compound Results']

        # Color code validation status
        status_colors = {
            'VALIDATED': 'C6EFCE',  # Green
            'LIKELY_SUCCESS': 'D9EAD3',  # Light green
            'UNCERTAIN': 'FFF2CC',  # Yellow
            'LIKELY_FAILURE': 'FCE5CD',  # Orange
            'FAILED': 'F4CCCC'  # Red
        }

        # Find status column
        status_col = None
        for col_idx, cell in enumerate(ws_compounds[1], 1):
            if cell.value == 'validation_status':
                status_col = col_idx
                break

        # Apply conditional formatting
        if status_col:
            col_letter = get_column_letter(status_col)
            for row in range(2, ws_compounds.max_row + 1):
                cell = ws_compounds[f'{col_letter}{row}']
                status = cell.value
                if status in status_colors:
                    fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type='solid'
                    )
                    cell.fill = fill

        # Bold headers
        for cell in ws_compounds[1]:
            cell.font = Font(bold=True)

        # Auto-size columns
        for column in ws_compounds.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_compounds.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws_compounds.freeze_panes = 'A2'

        # Save formatted workbook
        wb.save(output_path)

        return output_path
